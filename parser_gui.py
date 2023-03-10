import tkinter as tk
from tkinter.filedialog import askopenfilename
from tkinter.filedialog import askdirectory
import os
import argparse
from xml.etree import ElementTree
from dataclasses import dataclass, field
from typing import List
import json
# import pandas as pd
from openpyxl import Workbook
import os
import logging

logging.basicConfig(level=logging.INFO)

@dataclass
class Container:
    ecu_name: str = field(default_factory=str)
    config_value: str = field(default_factory=str)
    container_name: str = field(default_factory=str)
    definition_ref: str = field(default_factory=str)
    sub_container_name: str = field(default_factory=str)
    sub_definition_ref: str = field(default_factory=str)

class arxml_parsing:
        
    def __init__(self,arxml_path,db_path,xl_path) -> None:
        """Constructor for arxml_parsing"""
        self.arxml_path = arxml_path
        self.schema = "http://autosar.org/schema/r4.0"
        self.namespace = {"as":self.schema}
        self.root = ElementTree.parse(self.arxml_path)
        self.database: List[Container] = []
        self.extract_containers()
        self.get_database(db_path,xl_path)
        
    def extract_containers(self):
        ecu_name = self.root.find("as:AR-PACKAGES/as:AR-PACKAGE/as:SHORT-NAME",self.namespace).text
        ELEMENTS = "as:AR-PACKAGES/as:AR-PACKAGE/as:ELEMENTS/as:ECUC-MODULE-CONFIGURATION-VALUES"
        for ele in self.root.findall(ELEMENTS,self.namespace):
            config_value = next(iter(c.text  for c in ele if c.tag == f'{{{self.schema}}}SHORT-NAME'))
            CONTAINER_VALUE = "as:CONTAINERS/as:ECUC-CONTAINER-VALUE"
            for cont in ele.findall(CONTAINER_VALUE,self.namespace):
                container_name = next(iter(c.text  for c in cont if c.tag == f'{{{self.schema}}}SHORT-NAME'))
                definition_ref = next(iter(c.text  for c in cont if c.tag == f'{{{self.schema}}}DEFINITION-REF'))
                SUB_CONTAINERS = "as:SUB-CONTAINERS/as:ECUC-CONTAINER-VALUE"
                sub_cont = cont.find("as:SUB-CONTAINERS",self.namespace)
                if sub_cont:
                    for sub_cont in cont.findall(SUB_CONTAINERS,self.namespace):
                        sub_container_name = next(iter(c.text  for c in sub_cont if c.tag == f'{{{self.schema}}}SHORT-NAME'))
                        sub_definition_ref = next(iter(c.text  for c in sub_cont if c.tag == f'{{{self.schema}}}DEFINITION-REF')).split('/')[-1]
                        container = Container(ecu_name,config_value,container_name,definition_ref,sub_container_name,sub_definition_ref)
                        self.database.append(container)
                else:
                    container = Container(ecu_name,config_value,container_name,definition_ref,None,None)
                    self.database.append(container)
        logging.info("Containers and sub containers extracted")
    
    def get_database(self,db_path,xl_path):
        db = [cont.__dict__ for cont in self.database]
        # df = pd.DataFrame(db)
        with open(db_path,'w+') as f:
            json.dump(db,f,indent=4)
        # df.to_excel(xl_path+os.sep+'db.xlsx',index=False)
        keys = []
        wb = Workbook()
        ws = wb.active  
        for i in range(len(db)) :
            sub_obj = db[i]
            if i == 0 :
                keys = list(sub_obj.keys())
                for k in range(len(keys)) :
                    ws.cell(row = (i + 1), column = (k + 1), value = keys[k]);
            for j in range(len(keys)) :
                ws.cell(row = (i + 2), column = (j + 1), value = sub_obj[keys[j]]);
        wb.save(xl_path+os.sep+'db.xlsx')
        logging.info("Containers Database Excel Created")

class gui:

    def __init__(self) -> None:
        self.root = tk.Tk()
        self.root.geometry('250x125')
        self.filename = ''
        self.foldername = ''
        self.show()

    def browsefile(self):
        tk.Tk().withdraw()
        self.filename = askopenfilename()
        if os.path.isfile(self.filename) and "xml" in self.filename:
            logging.info("ARXML File selected")
            tk.messagebox.showinfo(title='ARXML File selected', message=self.filename )
        else:
            logging.error("ARXMl path not valid")
            tk.messagebox.showerror("showerror", "Please Update valid ARXMl path")
            

    def browsefolder(self):
        tk.Tk().withdraw()
        self.foldername = askdirectory()
        if os.path.isdir(self.foldername):
            logging.info("Excel directory selected")
            tk.messagebox.showinfo(title='Excel directory selected', message=self.foldername )
        else:
            logging.error("Excel path not valid")
            tk.messagebox.showerror("showerror", "Please Update valid Excel path")
            
        
        
    def run(self):
        if os.path.isfile(self.filename) and "xml" in self.filename  and os.path.isdir(self.foldername):
            arxml_parsing(self.filename,r'db.json',self.foldername)
            self.root.quit()
        else:
            tk.messagebox.showerror("showerror", "Please Update Both ARXMl and Excel valid paths")
            logging.error("Please Update Both ARXMl and Excel valid paths")

    def show(self):
        row_1 = tk.Frame(self.root)
        arxml = tk.Label(row_1, text="Choose input ARXML file")
        arxml.pack(side=tk.LEFT)
        arxml_path = tk.Button(row_1, text="Browse", command=self.browsefile)
        arxml_path.pack(side=tk.RIGHT, padx=5, pady=5)
        row_1.pack(side=tk.TOP, fill=tk.X, padx=10, pady=0)

        row_2 = tk.Frame(self.root)
        xl = tk.Label(row_2, text="Choose excel output directory")
        xl.pack(side=tk.LEFT)
        xl_path = tk.Button(row_2, text="Browse", command=self.browsefolder)
        xl_path.pack(side=tk.RIGHT, padx=5, pady=5)
        row_2.pack(side=tk.TOP, fill=tk.X, padx=10, pady=0)

        row_3 = tk.Frame(self.root)
        run = tk.Button(row_3, text='Parse',command=self.run)
        run.pack(side=tk.LEFT, padx=40, pady=5)
        quit = tk.Button(row_3, text='Quit', command=self.root.quit)
        quit.pack(side=tk.RIGHT, padx=40, pady=5)
        row_3.pack(side=tk.TOP, fill=tk.X, padx=10, pady=0)
        self.root.mainloop()



if __name__ == '__main__':
    cmd_parser = argparse.ArgumentParser(allow_abbrev=False)
    cmd_parser.add_argument(
        "--arxml",
        type=str,
        nargs="?",
        const="AutosarFile.xml",
        help="Enter ARXML path"
    )
    cmd_parser.add_argument(
        "--excel",
        type=str,
        nargs="?",
        const="./",
        help="Enter Excel path"
    )
    args = cmd_parser.parse_args()
    if os.path.isfile(str(args.arxml)) and os.path.isdir(str(args.excel)):
        logging.info("Starting ARXML Parser tool")
        arxml_parsing(args.arxml,r'db.json',args.excel)
    else:   
        logging.info("Starting ARXML Parser tool GUI")
        gui()
    logging.info('DONE')
